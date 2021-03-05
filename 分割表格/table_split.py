import openpyxl, threading,os
import tkinter as tk
from tkinter import ttk,messagebox
from tkinter.filedialog import askopenfilename
from openpyxl.styles import Alignment
from openpyxl.styles import numbers
from datetime import datetime


# 老规矩，多线程任务，防止窗口卡死
def my_thread(func, *args):
    t = threading.Thread(target=func, args=args)
    t.setDaemon(True)
    t.start()


# 二十六进制转换成十进制
def cfh_to_d(s):
    sum = 0
    for i, j in enumerate(s[::-1]):
        sum += (ord(j) - ord('A') + 1) * 26 ** i
    return sum


# 十进制转换成二十六进制
def cd_to_h(num):
    sequence = list(map(lambda x: chr(x), range(ord('A'), ord('Z') + 1)))
    L = []
    if num > 25:
        while True:
            d = int(num / 26)
            remainder = num % 26
            if d <= 25:
                L.insert(0, sequence[remainder])
                L.insert(0, sequence[d - 1])
                break
            else:
                L.insert(0, sequence[remainder])
                num = d - 1
    else:
        L.append(sequence[num])

    return "".join(L)


def sele_path():
    path = askopenfilename(title='选择需要拆分的excel表格', filetypes=[('XLSX', '*.xlsx')])
    excel_path_entry.set(path)
    wb = openpyxl.load_workbook(path)

    sele_sheet['value']=wb.sheetnames
    sele_sheet.current(0)

    ws=wb[sele_sheet.get()]

    #装载列表的数组
    col_list = []
    # 获取表格的最大列数目返回到tk窗口供其选择
    for i in range(len(list(ws.columns))):
        col_list.append(cd_to_h(i))
    sele_col['value'] = col_list
    sele_col.current(0)

    def handler(event):
        sele_col['value'] = col_list
        sele_col.current(0)
    sele_sheet.bind('<<ComboboxSelected>>',handler)


def screen_col():
    colunm = sele_col.get()
    path=excel_path_entry.get()
    wb = openpyxl.load_workbook(path,data_only=True)
    ws=wb[sele_sheet.get()]

    col = sele_col.get()
    print(col)
    screen_list = []
    for t in range(len(list(ws.rows))):
        comp = ws[col + str(t + 1)].value
        if comp == None or comp == '#N/A':
            pass
        else:
            if comp not in screen_list:
                screen_list.append(comp)
                #print(comp)
    print(screen_list)
    multiple_sele_window=tk.Toplevel(tb_s_window)
    multiple_sele_window.title('选择')
    multiple_sele_window.geometry('180x200')
    multiple_sele_window.resizable(0,0)

    heading_frame=tk.Frame(multiple_sele_window)
    heading_frame.pack(fill=tk.X)

    screen_button=tk.Button(heading_frame,text='切割',height=1,width=5,bg='#87CEEB',fg='white',command=lambda :my_thread(screen_work))
    screen_button.pack(side=tk.LEFT)
    tk.Label(heading_frame,text='选项',width=120,anchor='center',relief=tk.GROOVE).pack(side=tk.LEFT)

    #画板区域
    canvas_frame=tk.Frame(multiple_sele_window)
    canvas_frame.pack(fill=tk.BOTH,expand='yes')

    canvas=tk.Canvas(canvas_frame,bg='blue',width=150,height=150,scrollregion=(0,0,100,100))
    canvas.pack(side=tk.LEFT,fill=tk.Y,expand='yes')

    scroll=tk.Scrollbar(canvas_frame,orient=tk.VERTICAL)
    scroll.config(command=canvas.yview)
    canvas.configure(yscrollcommand=scroll.set)
    scroll.pack(side=tk.RIGHT,fill=tk.Y)

    canvas.bind_all("<MouseWheel>",lambda event:canvas.yview_scroll(int(-1 * (event.delta / 120)), "units"))

    tv_frame=tk.Frame(canvas_frame)
    tvv_frame=canvas.create_window(0,0,window=tv_frame,anchor='nw')

    button_frame=tk.Frame(tv_frame)
    button_frame.pack(side=tk.LEFT,fill=tk.Y,expand='yes')

    all_buttonvar = tk.IntVar()
    all_button = tk.Checkbutton(button_frame, text='', variable=all_buttonvar, command=lambda: my_thread(select_all))
    all_button.pack()
    all_buttonvar.set(1)

    tree=ttk.Treeview(tv_frame,columns=('options'),height=2,show='headings')
    #列表设置
    tree.column('options',width=120,anchor='center')
    tree.heading('options',text='全选')
    tree.pack(side=tk.LEFT,fill=tk.Y)

    screen_dir = {}
    #插入选项
    for tt,data_1 in enumerate(screen_list):
        tv_item=tree.insert('',tt,values=(data_1),tags=('oddrow'))
        cb = tk.Checkbutton(button_frame, text='', variable=tk.IntVar())
        cb['command'] = lambda item=tv_item:select_button(item)
        cb.select()
        cb.pack()
        screen_dir[tv_item] = [cb]

    #tree.configure('oddrow',font='Arial')
    ttk.Style().configure('Treeview',rowheight=27)

    height=(len(tree.get_children())+1)*27
    canvas.itemconfigure(tvv_frame,height=height)
    multiple_sele_window.update()
    canvas.config(scrollregion=canvas.bbox('all'))

    def select_all():
        #全选按钮的回调函数
        for item,[button] in screen_dir.items():
            if all_buttonvar.get()==1:
                button.select()
                tree.item(item,tag='select')
            else:
                button.deselect()
                tree.item(item,tags='oddrow')

    def select_button(item):
        #多选按钮的回调
        button=screen_dir[item][0]
        button_value=button.getvar(button['variable'])
        if button_value=='1':
            tree.item(item,tags='select')
        else:
            tree.item(item,tags='oddrow')
        all_button_select()

    def all_button_select():
        #根据所有按钮变化，全选按钮状态
        #循环所有按钮，当有一个按钮没有打勾时，全选按钮取消打勾
        for [button] in screen_dir.values():
            button_value=button.getvar(button['variable'])
            if button_value=='0':
                all_buttonvar.set(0)
                break
        else:
            all_buttonvar.set(1)

    def screen_work():
        col_list=[]
        for bb,[button] in enumerate(screen_dir.values()):
            if button.getvar(button['variable'])=='1':
                col_list.append(screen_list[bb])
        print(col_list)
        wb = openpyxl.load_workbook(excel_path_entry.get(), data_only=True)
        ws=wb[sele_sheet.get()]

        for j in col_list:
            wp = openpyxl.Workbook()
            wd = wp.active
            sheet_wp = wp['Sheet']

            for tt in range(len(list(ws.columns))):
                if ws[cd_to_h(tt) + '1'].value != None:
                    wd[cd_to_h(tt) + '1'] = ws[cd_to_h(tt) + '1'].value
                    sheet_wp.column_dimensions[cd_to_h(tt)].width = 12
                    wd[cd_to_h(tt) + '1'].alignment = Alignment(horizontal='center', vertical='center')
                    if type(ws[cd_to_h(tt) + '1'].value)==datetime:
                        wd[cd_to_h(tt) + '1'].number_format = numbers.FORMAT_DATE_YYYYMMDD2
            num = 2
            for z in range(len(list(ws.rows))):
                comp = ws[sele_col.get()+str(z+1)].value
                if comp == j:
                    for zz in range(len(list(ws.columns))):
                        wd[cd_to_h(zz) + str(num)] = ws[cd_to_h(zz) + str(z + 1)].value
                        wd[cd_to_h(zz) + str(num)].alignment = Alignment(horizontal='center', vertical='center')
                        # wp.save(f'{j}.xlsx')
                    num += 1
            date=datetime.now().strftime('%m-%d %H-%M')
            if not os.path.exists(date):
                os.mkdir(date)
            wp.save(f'{date}/{j}.xlsx')
        # print('已完成')
        tk.messagebox.showinfo(title='提示', message=f'已完成\n保存在{date}目录下')




if __name__ == '__main__':
    tb_s_window = tk.Tk()
    tb_s_window.geometry('300x150')
    tb_s_window.title('表格分割')
    tb_s_window.resizable(0, 0)

    # 表格label
    tk.Label(tb_s_window, text='表格：').place(x=10,y=20)#.grid(row=0, column=0)
    # 表格路径输入框
    excel_path_entry = tk.StringVar()
    tk.Entry(tb_s_window, width=26,textvariable=excel_path_entry).place(x=50,y=20)#.grid(row=0, column=1)
    # 表格打开
    tk.Button(tb_s_window, text='打开',bg='#87CEEB', command=lambda: my_thread(sele_path)).place(x=250,y=15)#.grid(row=0, column=2)

    # 选定筛选工作表
    tk.Label(tb_s_window, text='选择工作表：').place(x=10,y=50)#grid(row=1, column=0)
    # 选定筛选工作表输入框
    sele_sheet = ttk.Combobox(tb_s_window,width=20)
    sele_sheet.place(x=100,y=50)#.grid(row=1, column=1)
    # 选定筛选列
    tk.Label(tb_s_window, text='选择筛选类：').place(x=10, y=80)  # grid(row=1, column=0)
    # 选定筛选列输入框
    sele_col = ttk.Combobox(tb_s_window, width=20)
    sele_col.place(x=100, y=80)

    # 再设置一个筛选按钮
    tk.Button(tb_s_window,width=20, text='筛选',bg='#87CEEB' ,command=lambda: my_thread(screen_col)).place(x=80,y=110)#.grid(row=1, column=2)

    # 筛选选项展示

    tb_s_window.mainloop()
