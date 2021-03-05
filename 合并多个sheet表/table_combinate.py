import openpyxl, threading, os, datetime,webbrowser
import tkinter as tk
from tkinter import ttk, messagebox
from tkinter.filedialog import askdirectory
from datetime import datetime
from copy import copy
from openpyxl.utils import get_column_letter


# 老规矩，多线程任务，防止窗口卡死
def my_thread(func, *args):
    t = threading.Thread(target=func, args=args)
    t.setDaemon(True)
    t.start()


# 选择文件夹路径
def sele_path():
    path = askdirectory(title='请选择需要合并的文件夹')
    print(path)
    excel_path_entry.set(path)
    file_list = os.listdir(path)
    screen_list = []
    for file in file_list:
        if file[-5:] == '.xlsx' and file[:2] != '~$':
            screen_list.append(file)
    if len(screen_list) == 0:
        tk.messagebox.showinfo(title='提示', message='该目录中没有可合成的excel文件')


def sele_xlsx():
    file_list_path = excel_path_entry.get()
    try:
        file_list = os.listdir(file_list_path)
    except FileNotFoundError:
        tk.messagebox.showinfo(title='提示', message='请点击选择按钮，选择路径或者输入目录路径')
        quit()
    screen_list = []
    for file in file_list:
        if file[-5:] == '.xlsx' and  file[:2]!='~$':
            screen_list.append(file)
    if len(screen_list)==0:
        quit()


    multiple_sele_window = tk.Toplevel(tb_com_window)
    multiple_sele_window.title('选择')
    multiple_sele_window.geometry('300x200')
    # multiple_sele_window.resizable(0,0)

    heading_frame = tk.Frame(multiple_sele_window)
    heading_frame.pack(fill=tk.X)

    screen_button = tk.Button(heading_frame, text='合并', height=1, width=5, bg='#87CEEB', fg='white',
                              command=lambda: my_thread(screen_work))
    screen_button.pack(side=tk.LEFT)
    tk.Label(heading_frame, text='选项', width=120, anchor='center', relief=tk.GROOVE).pack(side=tk.LEFT)

    # 画板区域
    canvas_frame = tk.Frame(multiple_sele_window)
    canvas_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand='yes')

    canvas = tk.Canvas(canvas_frame, bg='white', width=150, height=150, scrollregion=(0, 0, 100, 100))
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand='yes')

    scroll = tk.Scrollbar(canvas_frame, orient=tk.VERTICAL)
    scroll.config(command=canvas.yview)
    canvas.configure(yscrollcommand=scroll.set)
    scroll.pack(side=tk.RIGHT, fill=tk.Y)

    canvas.bind_all("<MouseWheel>", lambda event: canvas.yview_scroll(int(-1 * (event.delta / 120)), "units"))

    tv_frame = tk.Frame(canvas_frame)
    tvv_frame = canvas.create_window(0, 0, window=tv_frame, anchor='nw')

    button_frame = tk.Frame(tv_frame)
    button_frame.pack(side=tk.LEFT, fill=tk.Y, expand='yes')

    all_buttonvar = tk.IntVar()
    all_button = tk.Checkbutton(button_frame, text='', variable=all_buttonvar, command=lambda: my_thread(select_all))
    all_button.pack()
    all_buttonvar.set(1)

    tree = ttk.Treeview(tv_frame, columns=('options'), height=2, show='headings')
    # 列表设置
    tree.column('options', width=250, anchor='center')
    tree.heading('options', text='全选')
    tree.pack(side=tk.LEFT, fill=tk.Y)

    screen_dir = {}
    # 插入选项
    for tt, data_1 in enumerate(screen_list):
        tv_item = tree.insert('', tt, values=(data_1), tags=('oddrow'))
        cb = tk.Checkbutton(button_frame, text='', variable=tk.IntVar())
        cb['command'] = lambda item=tv_item: select_button(item)
        cb.select()
        cb.pack()
        screen_dir[tv_item] = [cb]

    # tree.configure('oddrow',font='Arial')
    ttk.Style().configure('Treeview', rowheight=27)

    height = (len(tree.get_children()) + 1) * 27
    canvas.itemconfigure(tvv_frame, height=height)
    multiple_sele_window.update()
    canvas.config(scrollregion=canvas.bbox('all'))

    def select_all():
        # 全选按钮的回调函数
        for item, [button] in screen_dir.items():
            if all_buttonvar.get() == 1:
                button.select()
                tree.item(item, tag='select')
            else:
                button.deselect()
                tree.item(item, tags='oddrow')

    def select_button(item):
        # 多选按钮的回调
        button = screen_dir[item][0]
        button_value = button.getvar(button['variable'])
        if button_value == '1':
            tree.item(item, tags='select')
        else:
            tree.item(item, tags='oddrow')
        all_button_select()

    def all_button_select():
        # 根据所有按钮变化，全选按钮状态
        # 循环所有按钮，当有一个按钮没有打勾时，全选按钮取消打勾
        for [button] in screen_dir.values():
            button_value = button.getvar(button['variable'])
            if button_value == '0':
                all_buttonvar.set(0)
                break
        else:
            all_buttonvar.set(1)

    # 进行合并表格主要函数
    def screen_work():
        #对选择的选项中进行最后列表整理
        main_list = []
        for bb, [button] in enumerate(screen_dir.values()):
            if button.getvar(button['variable']) == '1':
                main_list.append(screen_list[bb])
        print(main_list)
        # 创建时间获取
        date_time = datetime.now().strftime('%Y-%m-%d-%H-%M')
        # print(date_time)
        file_list_path = excel_path_entry.get()
        # print(file_list_path)
        # 创建一个新的excel文件，用于最后的保存和合并
        wb = openpyxl.Workbook()
        sheet_number = 0
        for i, file in enumerate(main_list):
            # print(i ,file)
            file_path = file_list_path + '/' + file
            # print(file_path)每个excel文件的单独路径
            wp = openpyxl.load_workbook(file_path, data_only=True)
            # 根据表里所有的sheet进行操作
            for j, sheet in enumerate(wp.worksheets):
                # 在新的表格中创建一个sheet表
                new_sheet = wb.create_sheet(file[:-5] + '-' + sheet.title, sheet_number)
                sheet_number += 1
                # 获取原sheet的最大行数和列数
                max_row = sheet.max_row
                max_col = sheet.max_column
                # print(sheet.title)
                # 开始进行数据复制粘贴
                for row in range(1, max_row + 1):
                    for col in range(1, max_col + 1):
                        if sheet.cell(row=row, column=col).value!=None:
                            new_sheet.cell(row=row, column=col).value = copy(sheet.cell(row=row, column=col).value)
                        if sheet.cell(row=row, column=col).has_style:
                            new_sheet.cell(row=row, column=col).font = copy(sheet.cell(row=row, column=col).font)
                            new_sheet.cell(row=row, column=col).fill = copy(sheet.cell(row=row, column=col).fill)
                            new_sheet.cell(row=row, column=col).border = copy(sheet.cell(row=row, column=col).border)
                            new_sheet.cell(row=row, column=col).number_format = copy(
                                sheet.cell(row=row, column=col).number_format)
                            new_sheet.cell(row=row, column=col).protection = copy(
                                sheet.cell(row=row, column=col).protection)
                            new_sheet.cell(row=row, column=col).alignment = copy(sheet.cell(row=row, column=col).alignment)
                # 列与行大小整理
                for row_r in range(1, max_row + 1):
                    new_sheet.row_dimensions[row_r].height =copy( sheet.row_dimensions[row_r].height)
                    #print(sheet.row_dimensions[row_r].height)
                for col_c in range(1, max_col + 1):
                    #print(get_column_letter(col_c))
                    new_sheet.column_dimensions[get_column_letter(col_c)].width = copy(sheet.column_dimensions[
                        get_column_letter(col_c)].width)

                # 整合合并单元格
                merge_list = list(sheet.merged_cells)
                # print(merge_list)
                if len(merge_list) > 0:
                    for m in range(len(merge_list)):
                        new_sheet.merge_cells(str(merge_list[m]))
            # wp.close()
        wb.remove(wb['Sheet'])
        wb.save(file_list_path+'/'+date_time+'.xlsx')
        ask_question=tk.messagebox.askquestion(title='提示', message=f'已完成\n保存于{date_time}.xlsx\n是否打开此表格')
        if ask_question=="yes":
            webbrowser.open(file_list_path+'/'+date_time+'.xlsx')
        multiple_sele_window.destroy()



if __name__ == '__main__':
    tb_com_window = tk.Tk()
    tb_com_window.geometry('300x150')
    tb_com_window.title('表格合并')
    tb_com_window.resizable(0, 0)

    # 表格label
    tk.Label(tb_com_window, text='表格文件夹：',font=('宋体',11)).place(x=10, y=20)  # .grid(row=0, column=0)
    # 表格路径输入框
    excel_path_entry = tk.StringVar()
    tk.Entry(tb_com_window, width=25, textvariable=excel_path_entry).place(x=100, y=20)  # .grid(row=0, column=1)
    # 表格打开
    tk.Button(tb_com_window, text='选择', width=20,bg='#87CEEB', command=lambda: my_thread(sele_path)).place(x=80,
                                                                                                  y=60)  # .grid(row=0, column=2)

    # # 选定筛选工作表
    # tk.Label(tb_com_window, text='选择工作表：').place(x=10, y=50)  # grid(row=1, column=0)
    # # 选定筛选工作表输入框
    # sele_sheet = ttk.Combobox(tb_com_window, width=20)
    # sele_sheet.place(x=100, y=50)  # .grid(row=1, column=1)
    # # 选定筛选列
    # tk.Label(tb_com_window, text='选择筛选类：').place(x=10, y=80)  # grid(row=1, column=0)
    # # 选定筛选列输入框
    # sele_col = ttk.Combobox(tb_com_window, width=20)
    # sele_col.place(x=100, y=80)

    # 再设置一个筛选按钮
    tk.Button(tb_com_window, width=20, text='筛选文件', bg='#87CEEB', command=lambda: my_thread(sele_xlsx)).place(x=80,
                                                                                                            y=110)  # .grid(row=1, column=2)

    # 筛选选项展示

    tb_com_window.mainloop()
